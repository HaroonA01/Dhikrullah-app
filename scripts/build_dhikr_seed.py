#!/usr/bin/env python3
"""Generate db/seed/dhikrs.json + copy renamed audio from the master Excel.

Content (spreadsheet + source audio) lives outside this repo. Point the script
at it with the DHIKRULLAH_CONTENT_DIR environment variable, which should contain:
  - 'Dhikrullah Content Final.xlsx'
  - 'Cleaned Audio Files/'  (source audio dir)

Outputs (relative to repo root):
  - db/seed/dhikrs.json   (overwritten)
  - assets/audio/<slug>.mp3 (copied + renamed)
  - prints the slug map for use in db/audioMap.ts

Run:
  DHIKRULLAH_CONTENT_DIR=/path/to/content python3 scripts/build_dhikr_seed.py
"""
from __future__ import annotations

import json
import os
import re
import shutil
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
CONTENT_DIR = Path(
    os.environ.get("DHIKRULLAH_CONTENT_DIR", REPO_ROOT / "content")
)
EXCEL_PATH = CONTENT_DIR / "Dhikrullah Content Final.xlsx"
AUDIO_SRC_DIR = CONTENT_DIR / "Cleaned Audio Files"
AUDIO_DST_DIR = REPO_ROOT / "assets" / "audio"
SEED_PATH = REPO_ROOT / "db" / "seed" / "dhikrs.json"

CONTENT_VERSION = 6

# Sheet name -> list of category ids it expands into.
# "5 Daily Prayers" fans out to fajr/dhuhr/asr/maghrib/isha.
# "5 Daily Prayers (Fajr)" rows are only appended to fajr.
SHEET_CATEGORY_MAP: dict[str, list[str]] = {
    "All Day": ["all_day"],
    "Witr": ["witr"],
    "Waking Up": ["waking_up"],
    "5 Daily Prayers": ["fajr", "dhuhr", "asr", "maghrib", "isha"],
    "Morning": ["morning"],
    "Evening": ["evening"],
    "Before Bed": ["before_bed"],
}

FAJR_ONLY_CATEGORY_LABEL = "5 Daily Prayers (Fajr)"


def slugify_audio(filename: str) -> str:
    """`Allahumma … asbahna ...  ilaykal-masir_cleaned.mp3` -> `allahumma-asbahna-ilaykal-masir.mp3`."""
    name = filename
    # strip trailing _cleaned.mp3 (case-insensitive)
    name = re.sub(r"_cleaned\.mp3$", "", name, flags=re.IGNORECASE)
    # ellipsis variants -> dash
    name = name.replace("…", "-")
    name = name.replace("...", "-")
    # lowercase
    name = name.lower()
    # any non-[a-z0-9-] -> dash
    name = re.sub(r"[^a-z0-9-]+", "-", name)
    # collapse repeats, trim
    name = re.sub(r"-+", "-", name).strip("-")
    return name + ".mp3"


def clean_cell(value):
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        return v if v else None
    return value


def main() -> None:
    assert EXCEL_PATH.exists(), f"missing: {EXCEL_PATH}"
    assert AUDIO_SRC_DIR.exists(), f"missing: {AUDIO_SRC_DIR}"
    AUDIO_DST_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Group all output dhikrs by category id, preserving row order.
    by_category: dict[str, list[dict]] = {}

    # Track unique audio filenames seen so we copy each once.
    slug_map: dict[str, str] = {}  # original filename -> slug

    for sheet_name, category_ids in SHEET_CATEGORY_MAP.items():
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        col_idx = {h: i for i, h in enumerate(headers)}

        for row_idx in range(2, ws.max_row + 1):
            cells = [ws.cell(row=row_idx, column=c + 1).value for c in range(len(headers))]
            row = {h: cells[col_idx[h]] for h in headers}

            cat_label = clean_cell(row.get("Category"))
            arabic = clean_cell(row.get("Arabic"))
            if not arabic:
                continue  # skip empty filler rows

            audio_filename_raw = clean_cell(row.get("Audio File"))
            audio_slug = None
            if audio_filename_raw:
                if audio_filename_raw not in slug_map:
                    slug_map[audio_filename_raw] = slugify_audio(audio_filename_raw)
                audio_slug = slug_map[audio_filename_raw]

            target = row.get("Count")
            try:
                target = int(target) if target is not None else 1
            except (TypeError, ValueError):
                target = 1

            payload = {
                "arabic": arabic,
                "transliteration": clean_cell(row.get("Transliteration")) or "",
                "translation": clean_cell(row.get("English")) or "",
                "target": target,
                "description": clean_cell(row.get("Hadith")),
                "reference": clean_cell(row.get("Reference")),
                "grade": clean_cell(row.get("Grade")),
                "audioFilename": audio_slug,
            }

            # Determine which categories this row goes into.
            if cat_label == FAJR_ONLY_CATEGORY_LABEL:
                targets = ["fajr"]
            else:
                targets = category_ids

            for cid in targets:
                by_category.setdefault(cid, []).append(payload)

    # Build the final list with deterministic ids + sortOrder.
    out_dhikrs: list[dict] = []
    summary: list[tuple[str, int]] = []
    for cid, rows in by_category.items():
        for i, payload in enumerate(rows):
            d = {
                "id": f"{cid}-{i + 1:02d}",
                "categoryId": cid,
                **payload,
                "sortOrder": i,
            }
            # enforce schema field order for git diff stability
            ordered = {
                "id": d["id"],
                "categoryId": d["categoryId"],
                "arabic": d["arabic"],
                "transliteration": d["transliteration"],
                "translation": d["translation"],
                "target": d["target"],
                "description": d["description"],
                "reference": d["reference"],
                "grade": d["grade"],
                "audioFilename": d["audioFilename"],
                "sortOrder": d["sortOrder"],
            }
            out_dhikrs.append(ordered)
        summary.append((cid, len(rows)))

    # Write seed JSON.
    SEED_PATH.parent.mkdir(parents=True, exist_ok=True)
    with SEED_PATH.open("w", encoding="utf-8") as f:
        json.dump(
            {"contentVersion": CONTENT_VERSION, "dhikrs": out_dhikrs},
            f,
            ensure_ascii=False,
            indent=2,
        )
        f.write("\n")

    # Copy referenced audio files (rename to slug).
    copied = 0
    skipped_existing = 0
    missing = []
    for original, slug in sorted(slug_map.items()):
        src = AUDIO_SRC_DIR / original
        dst = AUDIO_DST_DIR / slug
        if not src.exists():
            missing.append(original)
            continue
        if dst.exists():
            shutil.copyfile(src, dst)  # overwrite for re-runs
            skipped_existing += 1
        else:
            shutil.copyfile(src, dst)
            copied += 1

    print(f"Wrote {SEED_PATH.relative_to(REPO_ROOT)} with {len(out_dhikrs)} dhikrs")
    print(f"contentVersion: {CONTENT_VERSION}")
    print("Per-category counts:")
    for cid, n in summary:
        print(f"  {cid:<12} {n}")
    print(f"\nAudio: {len(slug_map)} unique files. Copied {copied} new, refreshed {skipped_existing}.")
    if missing:
        print(f"MISSING in {AUDIO_SRC_DIR}:")
        for m in missing:
            print(f"  - {m}")

    print("\nSlug map (for db/audioMap.ts):")
    for original, slug in sorted(slug_map.items(), key=lambda kv: kv[1]):
        print(f"  '{slug}': require('@/assets/audio/{slug}'),")


if __name__ == "__main__":
    main()
